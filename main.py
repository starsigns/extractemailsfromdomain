
import sys
import os
import asyncio
import aiohttp
import re
from datetime import datetime
from collections import defaultdict
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QLabel, QFileDialog,
    QProgressBar, QMessageBox, QTableWidget, QTableWidgetItem, QHeaderView, 
    QHBoxLayout, QTextEdit, QCheckBox, QSpinBox, QGroupBox, QComboBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer

EMAIL_REGEX = r'\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b'

# Obfuscated email patterns for detection and de-obfuscation
OBFUSCATED_PATTERNS = [
    # Common obfuscation patterns with their replacements (improved order and precision)
    (r'\b([a-zA-Z0-9._%+-]+)\s+at\s+([a-zA-Z0-9.-]+)\s+dot\s+([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    (r'\b([a-zA-Z0-9._%+-]+)\s*\[at\]\s*([a-zA-Z0-9.-]+)\s*\[dot\]\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    (r'\b([a-zA-Z0-9._%+-]+)\s*\(at\)\s*([a-zA-Z0-9.-]+)\s*\(dot\)\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    (r'\b([a-zA-Z0-9._%+-]+)\s*\{at\}\s*([a-zA-Z0-9.-]+)\s*\{dot\}\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    (r'\b([a-zA-Z0-9._%+-]+)\s*<at>\s*([a-zA-Z0-9.-]+)\s*<dot>\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    
    # HTML entity obfuscation
    (r'\b([a-zA-Z0-9._%+-]+)\s*&#64;\s*([a-zA-Z0-9.-]+)\s*&#46;\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    (r'\b([a-zA-Z0-9._%+-]+)\s*&\#64;\s*([a-zA-Z0-9.-]+)\s*&\#46;\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    (r'\b([a-zA-Z0-9._%+-]+)\s*&\#x40;\s*([a-zA-Z0-9.-]+)\s*&\#x2E;\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    
    # Spaced obfuscation (more specific)
    (r'\b([a-zA-Z0-9._%+-]+)\s+@\s+([a-zA-Z0-9.-]+)\s+\.\s+([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    
    # Underscore/dash obfuscation
    (r'\b([a-zA-Z0-9._%+-]+)\s*_at_\s*([a-zA-Z0-9.-]+)\s*_dot_\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    (r'\b([a-zA-Z0-9._%+-]+)\s*-at-\s*([a-zA-Z0-9.-]+)\s*-dot-\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    (r'\b([a-zA-Z0-9._%+-]+)\s*\+at\+\s*([a-zA-Z0-9.-]+)\s*\+dot\+\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    (r'\b([a-zA-Z0-9._%+-]+)\s*\*at\*\s*([a-zA-Z0-9.-]+)\s*\*dot\*\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
    
    # Multi-language obfuscation
    (r'\b([a-zA-Z0-9._%+-]+)\s+arroba\s+([a-zA-Z0-9.-]+)\s+punto\s+([a-zA-Z]{2,})\b', r'\1@\2.\3'),  # Spanish
    (r'\b([a-zA-Z0-9._%+-]+)\s+arobase\s+([a-zA-Z0-9.-]+)\s+point\s+([a-zA-Z]{2,})\b', r'\1@\2.\3'),  # French
]

# Additional simple obfuscation replacements
SIMPLE_OBFUSCATIONS = {
    ' at ': '@',
    ' AT ': '@',
    '[at]': '@',
    '[AT]': '@',
    '(at)': '@',
    '(AT)': '@',
    '{at}': '@',
    '{AT}': '@',
    '<at>': '@',
    '<AT>': '@',
    '_at_': '@',
    '-at-': '@',
    '+at+': '@',
    '*at*': '@',
    ' dot ': '.',
    ' DOT ': '.',
    '[dot]': '.',
    '[DOT]': '.',
    '(dot)': '.',
    '(DOT)': '.',
    '{dot}': '.',
    '{DOT}': '.',
    '<dot>': '.',
    '<DOT>': '.',
    '_dot_': '.',
    '-dot-': '.',
    '+dot+': '.',
    '*dot*': '.',
    '&#64;': '@',
    '&#46;': '.',
    '&#x40;': '@',
    '&#x2E;': '.',
    '&amp;#64;': '@',
    '&amp;#46;': '.',
}

# Invalid email patterns to exclude
INVALID_PATTERNS = [
    # File extensions and technical patterns
    r'.*\.(png|jpg|jpeg|gif|svg|webp|ico|css|js|pdf|doc|docx|zip|rar|exe)$',
    r'.*@(example\.com|test\.com|localhost|127\.0\.0\.1)',
    r'^[^@]*@[^@]*\.(png|jpg|jpeg|gif|svg|webp|ico|css|js|pdf)$',
    r'.*@.*\.(png|jpg|jpeg|gif|svg|webp|ico|css|js)$',
    
    # Generic/placeholder patterns (enhanced)
    r'^abc@.*',
    r'^test@.*',
    r'^demo@.*',
    r'^sample@.*',
    r'^placeholder@.*',
    r'^dummy@.*',
    r'^fake@.*',
    r'^example@.*',
    r'^noreply@.*',
    r'^no-reply@.*',
    r'^donotreply@.*',
    r'^do-not-reply@.*',
    r'.*@abc\..*',
    r'.*@xyz\..*',
    r'.*@test\..*',
    r'.*@demo\..*',
    r'.*@sample\..*',
    r'.*@placeholder\..*',
    r'.*@dummy\..*',
    r'.*@fake\..*',
    r'.*@123\..*',
    r'.*@456\..*',
    r'.*@789\..*',
    r'^(a|b|c|x|y|z|1|2|3)@(a|b|c|x|y|z|1|2|3)\..*',
    r'^[a-z]{1,3}@[a-z]{1,3}\.(com|net|org)$',  # Short generic patterns like abc@xyz.com
    
    # Common placeholder domains
    r'.*@(abc\.com|xyz\.com|test\.org|demo\.net)$',
    r'.*@(sample\.com|placeholder\.org|dummy\.net)$',
    r'.*@(foo\.com|bar\.com|baz\.com)$',
    r'.*@(lorem\.com|ipsum\.com|dolor\.com)$',
    
    # Sequential/pattern emails (more specific)
    r'^(user|admin|test)[0-9]+@.*',  # Only block if followed by numbers
    r'^[a-z]+[0-9]+@[a-z]+[0-9]+\..*',
    r'^(email|mail)[0-9]+@.*',  # Only with numbers
    r'^(test|demo|sample)(user|admin|[0-9])+@.*',  # test1@, demouser@, etc.
    
    # Technical/system emails
    r'.*@sentry.*',
    r'.*@.*sentry.*',
    r'.*@ovh\.net$',
    r'.*@.*\.ovhcloud\.com$',
    r'^abuse@.*',
    r'^u003e.*',
    r'.*@exemple\.',
    r'.*@email\.',
    r'.*@domain\.',
    r'.*@mail\.com$',
    r'.*@.*\.abc\.xyz$',
    r'.*@.*wixpress\.com$',
    r'.*@.*doctolib\.fr$',
    r'.*@.*doctolib\.com$',
    r'.*@.*\.mssante\.fr$',
    r'.*@.*\.apicrypt\.org$',
    r'.*exceptions\..*',
    r'.*\.ingest\.sentry\..*',
    r'^[a-f0-9]{32}@.*',
    r'^[a-f0-9]{40}@.*',
    r'.*@.*\.local\.fr$',
    r'.*@prestashop\.com$',
    r'.*@themeisle\.com$',
    r'.*@cal\.com$',
    r'.*@tally\.so$',
    r'.*@linkeo\.com$',
    r'.*@doe\.com$',
    r'.*@.*\.website\.com$',
    r'.*@yourwebsite\.com$',
    r'.*@.*gmail\.com$',  # Too many generic gmail addresses
    
    # Additional obvious fake patterns
    r'^(asdf|qwerty|12345|abcde)@.*',
    r'.*@(asdf|qwerty|12345|abcde)\..*',
    r'^[a-z]{1,2}@[a-z]{1,2}\.(co|me|us|io)$',  # Very short domains
]

class ExtractWorker(QThread):
    progress = pyqtSignal(int, int)
    result = pyqtSignal(list, dict)  # results, tld_counts
    status = pyqtSignal(str)
    auto_save = pyqtSignal()
    partial_results = pyqtSignal(list)
    speed_update = pyqtSignal(float, str)  # emails/min, time remaining

    def __init__(self, domains, auto_export_path=None):
        super().__init__()
        self.domains = domains
        self.results = []
        self.auto_export_path = auto_export_path
        self.processed_count = 0
        self.seen_emails = set()
        self.tld_counts = defaultdict(int)
        self.paused = False
        self.stopped = False
        self.start_time = None

    def pause(self):
        self.paused = True

    def resume(self):
        self.paused = False

    def stop(self):
        self.stopped = True
        self.paused = False

    def run(self):
        self.start_time = datetime.now()
        asyncio.run(self.async_run())

    async def async_run(self):
        connector = aiohttp.TCPConnector(limit=200, limit_per_host=20, ttl_dns_cache=300)
        timeout = aiohttp.ClientTimeout(total=3, connect=1)
        async with aiohttp.ClientSession(
            connector=connector, 
            timeout=timeout,
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        ) as session:
            # Process domains in larger batches for better performance
            batch_size = 100
            update_frequency = 2  # Update GUI every 2 batches to reduce hanging
            
            for i in range(0, len(self.domains), batch_size):
                # Check for pause/stop
                while self.paused and not self.stopped:
                    await asyncio.sleep(0.1)
                
                if self.stopped:
                    break
                
                batch = self.domains[i:i+batch_size]
                # Process batch domains concurrently
                tasks = [self.process_domain(session, domain) for domain in batch]
                await asyncio.gather(*tasks, return_exceptions=True)
                
                current_progress = min(i + batch_size, len(self.domains))
                
                # Calculate speed and ETA
                if self.start_time:
                    elapsed = (datetime.now() - self.start_time).total_seconds()
                    if elapsed > 0:
                        emails_per_sec = len(self.results) / elapsed
                        emails_per_min = emails_per_sec * 60
                        
                        remaining_domains = len(self.domains) - current_progress
                        if current_progress > 0:
                            avg_time_per_domain = elapsed / current_progress
                            eta_seconds = remaining_domains * avg_time_per_domain
                            eta_str = f"{int(eta_seconds//3600)}h {int((eta_seconds%3600)//60)}m"
                        else:
                            eta_str = "Calculating..."
                        
                        self.speed_update.emit(emails_per_min, eta_str)
                
                # Only update GUI every few batches to prevent hanging
                if (i // batch_size) % update_frequency == 0 or current_progress >= len(self.domains):
                    self.progress.emit(current_progress, len(self.domains))
                    self.partial_results.emit(self.results.copy())
                
                # Auto-save every 500 processed domains (less frequent for speed)
                if self.auto_export_path and current_progress % 500 == 0:
                    self.auto_save.emit()
                    
            if not self.stopped:
                self.result.emit(self.results, dict(self.tld_counts))
                self.status.emit('Extraction complete!')
            else:
                self.status.emit('Extraction stopped by user')

    async def process_domain(self, session, domain):
        url = domain if domain.startswith('http') else f'http://{domain}'
        domain_emails = set()  # Track emails for this specific domain
        try:
            html = await self.fetch(session, url)
            if html:
                # Extract emails from home page
                home_emails = self.extract_emails(html)
                for email in home_emails:
                    email_domain_key = f"{email}|{domain}"
                    if email_domain_key not in self.seen_emails:
                        self.seen_emails.add(email_domain_key)
                        self.results.append({'domain': domain, 'email': email, 'source_url': url})
                        domain_emails.add(email)
                        
                        # Track TLD
                        email_domain = email.split('@')[1].lower()
                        if '.' in email_domain:
                            tld = email_domain.split('.')[-1]
                            self.tld_counts[tld] += 1
                
                # Extract and process only priority links (faster filtering)
                links = self.extract_priority_links(html, url)
                limited_links = list(links)[:5]  # Reduced from 10 to 5 for speed
                
                if limited_links:
                    tasks = [self.fetch_and_extract_emails_with_source(session, link) for link in limited_links]
                    results = await asyncio.gather(*tasks, return_exceptions=True)
                    for result in results:
                        if isinstance(result, tuple) and len(result) == 2:
                            link_emails, source_url = result
                            for email in link_emails:
                                email_domain_key = f"{email}|{domain}"
                                if email_domain_key not in self.seen_emails:
                                    self.seen_emails.add(email_domain_key)
                                    self.results.append({'domain': domain, 'email': email, 'source_url': source_url})
                                    domain_emails.add(email)
                                    
                                    # Track TLD
                                    email_domain = email.split('@')[1].lower()
                                    if '.' in email_domain:
                                        tld = email_domain.split('.')[-1]
                                        self.tld_counts[tld] += 1
        except Exception:
            pass

    async def fetch(self, session, url):
        try:
            async with session.get(url, timeout=3, allow_redirects=False) as resp:
                if resp.status == 200:
                    # Limit response size for speed (1MB max)
                    content = await resp.read()
                    if len(content) > 1024*1024:  # 1MB limit
                        content = content[:1024*1024]
                    return content.decode('utf-8', errors='ignore')
        except Exception:
            return None
        return None

    async def fetch_and_extract_emails_with_source(self, session, url):
        html = await self.fetch(session, url)
        if html:
            return self.extract_emails(html), url
        return set(), url

    def extract_emails(self, html):
        # Use faster regex with compiled pattern for regular emails
        if not hasattr(self, '_email_pattern'):
            self._email_pattern = re.compile(EMAIL_REGEX, re.IGNORECASE)
        
        valid_emails = set()
        
        # 1. Extract regular emails first
        raw_emails = set(self._email_pattern.findall(html))
        for email in raw_emails:
            email = email.lower().strip()
            if self.is_valid_email(email):
                valid_emails.add(email)
        
        # 2. Extract obfuscated emails
        obfuscated_emails = self.extract_obfuscated_emails(html)
        for email in obfuscated_emails:
            if self.is_valid_email(email):
                valid_emails.add(email)
        
        # Limit emails per page for speed
        return set(list(valid_emails)[:25])  # Increased from 20 to 25 due to obfuscation

    def extract_obfuscated_emails(self, text):
        """Extract and de-obfuscate hidden email addresses"""
        obfuscated_emails = set()
        
        # 1. First, try complex patterns on original text (before simple replacements)
        for pattern, replacement in OBFUSCATED_PATTERNS:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                if isinstance(match, tuple) and len(match) >= 3:
                    # Reconstruct email from captured groups
                    email = f"{match[0].strip()}@{match[1].strip()}.{match[2].strip()}"
                    # Validate the reconstructed email
                    if re.match(EMAIL_REGEX, email, re.IGNORECASE):
                        obfuscated_emails.add(email.lower())
        
        # 2. Apply targeted simple replacements (only in email-like contexts)
        working_text = text
        
        # First, let's look for email-like patterns and replace them carefully
        # Pattern: word + obfuscated @ + word + obfuscated . + word
        email_context_patterns = [
            (r'\b([a-zA-Z0-9._%+-]+)\s*\[at\]\s*([a-zA-Z0-9.-]+)\s*\[dot\]\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
            (r'\b([a-zA-Z0-9._%+-]+)\s*\(at\)\s*([a-zA-Z0-9.-]+)\s*\(dot\)\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
            (r'\b([a-zA-Z0-9._%+-]+)\s*\{at\}\s*([a-zA-Z0-9.-]+)\s*\{dot\}\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
            (r'\b([a-zA-Z0-9._%+-]+)\s*<at>\s*([a-zA-Z0-9.-]+)\s*<dot>\s*([a-zA-Z]{2,})\b', r'\1@\2.\3'),
        ]
        
        for pattern, replacement in email_context_patterns:
            working_text = re.sub(pattern, replacement, working_text, flags=re.IGNORECASE)
        
        # Handle HTML entities
        working_text = working_text.replace('&#64;', '@').replace('&#46;', '.')
        working_text = working_text.replace('&#x40;', '@').replace('&#x2E;', '.')
        
        # 3. Look for newly formed emails after targeted replacements
        if not hasattr(self, '_email_pattern'):
            self._email_pattern = re.compile(EMAIL_REGEX, re.IGNORECASE)
        
        new_emails = set(self._email_pattern.findall(working_text))
        for email in new_emails:
            email = email.lower().strip()
            # Only add if it passes validation and isn't already found
            if email not in obfuscated_emails and re.match(EMAIL_REGEX, email, re.IGNORECASE):
                obfuscated_emails.add(email)
        
        # 4. Handle special cases like base64 encoded or other encodings
        obfuscated_emails.update(self.extract_encoded_emails(text))
        
        return obfuscated_emails

    def extract_encoded_emails(self, text):
        """Extract emails from encoded content (base64, URL encoding, etc.)"""
        encoded_emails = set()
        
        # Look for base64 encoded emails (common obfuscation)
        import base64
        
        # Find potential base64 strings that might contain emails
        base64_pattern = r'[A-Za-z0-9+/]{20,}={0,2}'
        base64_matches = re.findall(base64_pattern, text)
        
        for match in base64_matches:
            try:
                # Attempt to decode
                decoded = base64.b64decode(match).decode('utf-8', errors='ignore')
                # Look for emails in decoded content
                if '@' in decoded and '.' in decoded:
                    potential_emails = re.findall(EMAIL_REGEX, decoded, re.IGNORECASE)
                    for email in potential_emails:
                        encoded_emails.add(email.lower().strip())
            except:
                continue
        
        # Look for URL encoded emails
        import urllib.parse
        url_encoded_pattern = r'[a-zA-Z0-9%]{10,}'
        url_matches = re.findall(url_encoded_pattern, text)
        
        for match in url_matches:
            if '%' in match:
                try:
                    decoded = urllib.parse.unquote(match)
                    if '@' in decoded and '.' in decoded:
                        potential_emails = re.findall(EMAIL_REGEX, decoded, re.IGNORECASE)
                        for email in potential_emails:
                            encoded_emails.add(email.lower().strip())
                except:
                    continue
        
        return encoded_emails

    def is_valid_email(self, email):
        # Check against invalid patterns
        for pattern in INVALID_PATTERNS:
            if re.match(pattern, email, re.IGNORECASE):
                return False
        
        # Additional validation checks
        if len(email) < 5 or len(email) > 254:
            return False
        
        # Check for common invalid patterns
        if email.count('@') != 1:
            return False
        
        local, domain = email.split('@')
        
        # Local part validation
        if len(local) == 0 or len(local) > 64:
            return False
        
        # Domain validation
        if len(domain) == 0 or len(domain) > 253:
            return False
        
        # Check if domain has at least one dot
        if '.' not in domain:
            return False
        
        # Check for common spam/invalid domains and patterns (be more specific)
        spam_domains = [
            'noreply', 'no-reply', 'donotreply', 'example.com', 'test.com',
            'sentry', 'ovh.net', 'ovhcloud.com', 'wixpress.com', 'doctolib',
            'mssante.fr', 'apicrypt.org', 'prestashop.com', 'themeisle.com',
            'linkeo.com', 'tally.so', 'cal.com'
        ]
        # Only flag as spam if domain exactly matches or contains these specific patterns
        if (domain.lower() in spam_domains or 
            any(spam in domain.lower() and len(spam) > 3 for spam in spam_domains)):
            return False
        
        # Check for hash-like patterns (32+ hex characters)
        if re.match(r'^[a-f0-9]{32,}@', email):
            return False
        
        # Check for obvious test emails
        test_patterns = [
            'exemple', 'example', 'test@', 'demo@', 'sample@',
            'jean.dupont', 'marie.durand', 'martin.durand', 'john@doe',
            'nom@', 'votre@', 'your@', 'email@domain', 'user@mail'
        ]
        if any(test in email.lower() for test in test_patterns):
            return False
        
        # Enhanced check for generic/placeholder patterns
        local_lower = local.lower()
        domain_lower = domain.lower()
        
        # Check for obvious placeholders (be more selective with 'info' and 'admin')
        obvious_placeholders = ['abc', 'xyz', 'test', 'demo', 'sample', 'placeholder', 'dummy', 'fake']
        generic_domains = ['abc.com', 'xyz.com', 'test.com', 'demo.com', 'sample.com', 'placeholder.com']
        
        if local_lower in obvious_placeholders or domain_lower in generic_domains:
            return False
        
        # Check for very short local/domain combinations (likely placeholders)
        if len(local) <= 3 and len(domain.split('.')[0]) <= 3:
            return False
        
        # Check for sequential patterns (a@b.com, 1@2.com, etc.)
        if (len(local) == 1 and len(domain.split('.')[0]) == 1 and 
            local.isalpha() and domain.split('.')[0].isalpha()):
            return False
        
        # Check for common keyboard patterns
        keyboard_patterns = ['qwerty', 'asdf', 'zxcv', '12345', 'abcde']
        if any(pattern in local_lower or pattern in domain_lower for pattern in keyboard_patterns):
            return False
        
        # Check if local part is just repeated characters
        if len(set(local_lower)) == 1 and len(local) > 1:  # aaaa@domain.com
            return False
        
        # Check for info/admin only on obviously fake domains
        if local_lower in ['info', 'admin'] and any(fake in domain_lower for fake in ['test', 'demo', 'sample', 'fake', 'dummy']):
            return False
        
        # Check for u003e prefix (encoded characters)
        if email.startswith('u003e'):
            return False
        
        return True

    def extract_priority_links(self, html, base_url):
        """Extract only high-priority links likely to contain emails"""
        soup = BeautifulSoup(html, 'html.parser')
        links = set()
        
        # Priority keywords for email-likely pages
        priority_keywords = [
            'contact', 'about', 'team', 'staff', 'management', 'direction',
            'legal', 'mentions', 'imprint', 'impressum', 'support', 'help'
        ]
        
        for a in soup.find_all('a', href=True):
            href = a['href'].lower()
            link_text = a.get_text().lower()
            
            # Check if href or link text contains priority keywords
            if any(keyword in href or keyword in link_text for keyword in priority_keywords):
                if href.startswith('http'):
                    links.add(a['href'])
                elif href.startswith('/'):
                    links.add(self.join_url(base_url, a['href']))
                    
        return links

    def extract_links(self, html, base_url):
        soup = BeautifulSoup(html, 'html.parser')
        links = set()
        for a in soup.find_all('a', href=True):
            href = a['href']
            if href.startswith('http'):
                links.add(href)
            elif href.startswith('/'):
                links.add(self.join_url(base_url, href))
        return links

    def join_url(self, base, path):
        if base.endswith('/'):
            base = base[:-1]
        if path.startswith('/'):
            path = path[1:]
        return f'{base}/{path}'

class EmailExtractorApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Email Extractor Pro - Advanced Domain Email Harvester')
        self.resize(1000, 650)  # More compact height
        self.domains = []
        self.results = []
        self.tld_counts = {}
        self.worker = None
        self.pending_results = []
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.process_pending_updates)
        self.update_timer.setSingleShot(True)
        self.exclude_patterns = []
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(12, 12, 12, 12)

        # Modern title with gradient background
        title = QLabel('🚀 Email Extractor Pro')
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("""
            QLabel {
                font-size: 22px;
                font-weight: bold;
                color: white;
                margin: 8px 0;
                padding: 12px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                            stop:0 #667eea, stop:1 #764ba2);
                border-radius: 10px;
                border: none;
            }
        """)
        layout.addWidget(title)

        # Compact file selection with modern styling
        file_group = QGroupBox("📁 Domain File")
        file_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin: 3px 0;
                padding-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px;
                color: #3498db;
            }
        """)
        file_layout = QVBoxLayout()
        file_layout.setSpacing(6)
        
        # File selection row
        file_row = QHBoxLayout()
        self.file_label = QLabel('No file selected')
        self.file_label.setStyleSheet("""
            QLabel {
                padding: 6px 10px;
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                color: #6c757d;
                font-size: 11px;
            }
        """)
        
        self.browse_btn = QPushButton('📂 Browse')
        self.browse_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
                min-width: 70px;
            }
            QPushButton:hover { background-color: #2980b9; }
            QPushButton:pressed { background-color: #21618c; }
        """)
        self.browse_btn.clicked.connect(self.browse_file)
        
        self.validate_btn = QPushButton('✓ Validate')
        self.validate_btn.setEnabled(False)
        self.validate_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
                min-width: 70px;
            }
            QPushButton:hover { background-color: #218838; }
            QPushButton:disabled { background-color: #6c757d; }
        """)
        self.validate_btn.clicked.connect(self.validate_domains)
        
        file_row.addWidget(self.file_label, 3)
        file_row.addWidget(self.browse_btn)
        file_row.addWidget(self.validate_btn)
        file_layout.addLayout(file_row)
        
        # Exclude patterns (compact)
        exclude_row = QHBoxLayout()
        exclude_label = QLabel('Exclude:')
        exclude_label.setStyleSheet("font-size: 11px; color: #6c757d; min-width: 50px;")
        self.exclude_input = QTextEdit()
        self.exclude_input.setMaximumHeight(35)
        self.exclude_input.setStyleSheet("""
            QTextEdit {
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 4px;
                font-size: 10px;
                background-color: #ffffff;
            }
        """)
        self.exclude_input.setPlaceholderText('Enter patterns to exclude (one per line)')
        exclude_row.addWidget(exclude_label)
        exclude_row.addWidget(self.exclude_input)
        file_layout.addLayout(exclude_row)
        
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # Side-by-side validation results and TLD stats
        stats_row = QHBoxLayout()
        stats_row.setSpacing(8)
        
        # Domain validation results (left side)
        validation_group = QGroupBox("📊 Validation Results")
        validation_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #28a745;
                border-radius: 8px;
                margin: 3px 0;
                padding-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px;
                color: #28a745;
            }
        """)
        validation_layout = QVBoxLayout()
        self.domain_info = QLabel('')
        self.domain_info.setStyleSheet("""
            QLabel {
                font-size: 12px;
                font-weight: 500;
                color: #2c3e50;
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
                min-height: 110px;
                line-height: 1.4;
            }
        """)
        validation_layout.addWidget(self.domain_info)
        validation_group.setLayout(validation_layout)
        
        # TLD distribution (right side)
        tld_group = QGroupBox("🌐 TLD Distribution")
        tld_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #fd7e14;
                border-radius: 8px;
                margin: 3px 0;
                padding-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px;
                color: #fd7e14;
            }
        """)
        tld_layout = QVBoxLayout()
        self.tld_info = QLabel('Load domains to see TLD stats')
        self.tld_info.setStyleSheet("""
            QLabel {
                font-size: 11px;
                font-weight: 500;
                font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
                color: #2c3e50;
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
                min-height: 110px;
                line-height: 1.3;
            }
        """)
        self.tld_info.setWordWrap(True)
        tld_layout.addWidget(self.tld_info)
        tld_group.setLayout(tld_layout)
        
        stats_row.addWidget(validation_group)
        stats_row.addWidget(tld_group)
        layout.addLayout(stats_row)

        # Compact control panel
        control_group = QGroupBox("🎮 Controls")
        control_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #6f42c1;
                border-radius: 8px;
                margin: 3px 0;
                padding-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px;
                color: #6f42c1;
            }
        """)
        control_layout = QHBoxLayout()
        control_layout.setSpacing(8)
        
        button_style = """
            QPushButton {
                border: none;
                padding: 10px 15px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 11px;
                min-width: 70px;
            }
        """
        
        self.start_btn = QPushButton('▶️ Start')
        self.start_btn.setEnabled(False)
        self.start_btn.setStyleSheet(button_style + """
            QPushButton {
                background-color: #28a745;
                color: white;
            }
            QPushButton:hover { background-color: #218838; }
            QPushButton:disabled { background-color: #6c757d; }
        """)
        self.start_btn.clicked.connect(self.start_extraction)

        self.pause_btn = QPushButton('⏸️ Pause')
        self.pause_btn.setEnabled(False)
        self.pause_btn.setStyleSheet(button_style + """
            QPushButton {
                background-color: #ffc107;
                color: #212529;
            }
            QPushButton:hover { background-color: #e0a800; }
            QPushButton:disabled { background-color: #6c757d; }
        """)
        self.pause_btn.clicked.connect(self.pause_extraction)

        self.resume_btn = QPushButton('▶️ Resume')
        self.resume_btn.setEnabled(False)
        self.resume_btn.setStyleSheet(button_style + """
            QPushButton {
                background-color: #17a2b8;
                color: white;
            }
            QPushButton:hover { background-color: #138496; }
            QPushButton:disabled { background-color: #6c757d; }
        """)
        self.resume_btn.clicked.connect(self.resume_extraction)

        self.stop_btn = QPushButton('⏹️ Stop')
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet(button_style + """
            QPushButton {
                background-color: #dc3545;
                color: white;
            }
            QPushButton:hover { background-color: #c82333; }
            QPushButton:disabled { background-color: #6c757d; }
        """)
        self.stop_btn.clicked.connect(self.stop_extraction)
        
        control_layout.addWidget(self.start_btn)
        control_layout.addWidget(self.pause_btn)
        control_layout.addWidget(self.resume_btn)
        control_layout.addWidget(self.stop_btn)
        control_layout.addStretch()
        control_group.setLayout(control_layout)
        layout.addWidget(control_group)

        # Compact progress section
        progress_group = QGroupBox("📈 Progress")
        progress_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #e83e8c;
                border-radius: 8px;
                margin: 3px 0;
                padding-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px;
                color: #e83e8c;
            }
        """)
        progress_layout = QVBoxLayout()
        progress_layout.setSpacing(4)
        
        self.progress = QProgressBar()
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #ced4da;
                border-radius: 5px;
                text-align: center;
                font-weight: bold;
                font-size: 10px;
                height: 18px;
            }
            QProgressBar::chunk {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                stop:0 #667eea, stop:1 #764ba2);
                border-radius: 3px;
            }
        """)
        progress_layout.addWidget(self.progress)

        status_row = QHBoxLayout()
        self.status = QLabel('')
        self.status.setStyleSheet("font-size: 11px; color: #495057;")
        self.speed_label = QLabel('')
        self.speed_label.setStyleSheet("font-size: 11px; color: #495057;")
        status_row.addWidget(self.status)
        status_row.addStretch()
        status_row.addWidget(self.speed_label)
        progress_layout.addLayout(status_row)
        
        progress_group.setLayout(progress_layout)
        layout.addWidget(progress_group)

        # Compact results section
        results_group = QGroupBox("📧 Results")
        results_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #20c997;
                border-radius: 8px;
                margin: 3px 0;
                padding-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 8px;
                padding: 0 4px;
                color: #20c997;
            }
        """)
        results_layout = QVBoxLayout()
        results_layout.setSpacing(6)
        
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(['Domain', 'Email', 'Source URL'])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #dee2e6;
                background-color: white;
                alternate-background-color: #f8f9fa;
                font-size: 10px;
            }
            QHeaderView::section {
                background-color: #e9ecef;
                color: #495057;
                padding: 6px;
                border: 1px solid #dee2e6;
                font-weight: bold;
                font-size: 10px;
            }
        """)
        self.table.setAlternatingRowColors(True)
        results_layout.addWidget(self.table)

        # Compact export section
        export_layout = QHBoxLayout()
        export_layout.setSpacing(8)
        
        self.export_excel_btn = QPushButton('📊 Export Excel')
        self.export_excel_btn.setEnabled(False)
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #198754;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
                min-width: 90px;
            }
            QPushButton:hover { background-color: #157347; }
            QPushButton:disabled { background-color: #6c757d; }
        """)
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        
        self.export_txt_btn = QPushButton('📄 Export Text')
        self.export_txt_btn.setEnabled(False)
        self.export_txt_btn.setStyleSheet("""
            QPushButton {
                background-color: #6f42c1;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11px;
                min-width: 90px;
            }
            QPushButton:hover { background-color: #5a32a3; }
            QPushButton:disabled { background-color: #6c757d; }
        """)
        self.export_txt_btn.clicked.connect(self.export_to_text)
        
        export_layout.addWidget(self.export_excel_btn)
        export_layout.addWidget(self.export_txt_btn)
        export_layout.addStretch()
        
        results_layout.addLayout(export_layout)
        results_group.setLayout(results_layout)
        layout.addWidget(results_group)

        self.setLayout(layout)

    def validate_domain(self, domain):
        """Validate a single domain format"""
        domain = domain.strip().lower()
        if not domain:
            return False, "Empty domain"
        
        # Remove protocol if present
        if domain.startswith(('http://', 'https://')):
            domain = domain.split('://', 1)[1]
        
        # Remove path if present
        if '/' in domain:
            domain = domain.split('/')[0]
        
        # Basic domain validation
        if not re.match(r'^[a-zA-Z0-9][a-zA-Z0-9.-]*[a-zA-Z0-9]$', domain):
            return False, "Invalid format"
        
        if '..' in domain or domain.startswith('.') or domain.endswith('.'):
            return False, "Invalid format"
        
        if '.' not in domain:
            return False, "Missing TLD"
        
        return True, domain

    def validate_domains(self):
        """Validate and clean domain list"""
        if not self.domains:
            return
        
        valid_domains = []
        invalid_domains = []
        duplicates = set()
        seen = set()
        
        # Get exclude patterns
        exclude_text = self.exclude_input.toPlainText().strip()
        self.exclude_patterns = [p.strip().lower() for p in exclude_text.split('\n') if p.strip()]
        
        for domain in self.domains:
            is_valid, result = self.validate_domain(domain)
            
            if is_valid:
                clean_domain = result
                
                # Check against exclude patterns
                excluded = False
                for pattern in self.exclude_patterns:
                    if pattern in clean_domain:
                        excluded = True
                        break
                
                if excluded:
                    invalid_domains.append(f"{domain} (excluded)")
                elif clean_domain in seen:
                    duplicates.add(clean_domain)
                else:
                    valid_domains.append(clean_domain)
                    seen.add(clean_domain)
            else:
                invalid_domains.append(f"{domain} ({result})")
        
        self.domains = valid_domains
        
        # Show validation results
        total_original = len(self.domains) + len(invalid_domains) + len(duplicates)
        valid_count = len(valid_domains)
        invalid_count = len(invalid_domains)
        duplicate_count = len(duplicates)
        
        # Categorize by TLD
        tld_stats = {}
        for domain in valid_domains:
            tld = domain.split('.')[-1]
            tld_stats[tld] = tld_stats.get(tld, 0) + 1
        
        # Sort TLDs by count
        top_tlds = sorted(tld_stats.items(), key=lambda x: x[1], reverse=True)[:10]
        
        info_text = f"""✅ Valid: {valid_count:,}
❌ Invalid/Excluded: {invalid_count:,}
🔄 Duplicates removed: {duplicate_count:,}
📊 Total processed: {total_original:,}

⏱️ Est. time: {self.estimate_processing_time(valid_count)}"""
        
        self.domain_info.setText(info_text)
        
        # Update TLD info in a compact, multi-column format
        if top_tlds:
            # Split TLDs into two columns for better space utilization
            mid_point = (len(top_tlds) + 1) // 2
            left_column = top_tlds[:mid_point]
            right_column = top_tlds[mid_point:]
            
            tld_lines = []
            for i in range(max(len(left_column), len(right_column))):
                left_item = f".{left_column[i][0]}: {left_column[i][1]:,}" if i < len(left_column) else ""
                right_item = f".{right_column[i][0]}: {right_column[i][1]:,}" if i < len(right_column) else ""
                
                # Format line with proper spacing (adjust spacing for better alignment)
                if left_item and right_item:
                    line = f"{left_item:<18} {right_item}"
                elif left_item:
                    line = left_item
                else:
                    line = f"{'':18} {right_item}"
                
                tld_lines.append(line)
            
            tld_text = "Top TLD Extensions:\n\n" + "\n".join(tld_lines)
        else:
            tld_text = "Load and validate domains\nto see TLD statistics"
        
        self.tld_info.setText(tld_text)
        
        if invalid_count > 0 or duplicate_count > 0:
            reply = QMessageBox.question(
                self, 'Validation Results',
                f"Found {invalid_count} invalid and {duplicate_count} duplicate domains.\n"
                f"Continue with {valid_count} valid domains?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.No:
                return
        
        self.start_btn.setEnabled(len(self.domains) > 0)

    def estimate_processing_time(self, domain_count):
        """Estimate processing time based on domain count"""
        # Rough estimate: ~3-5 seconds per domain with current optimizations
        avg_time_per_domain = 4  # seconds
        total_seconds = domain_count * avg_time_per_domain
        
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        
        if hours > 0:
            return f"~{hours}h {minutes}m"
        else:
            return f"~{minutes}m"

    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Open file', '', 'Text Files (*.txt)')
        if file_path:
            self.file_label.setText(os.path.basename(file_path))
            with open(file_path, 'r', encoding='utf-8') as f:
                self.domains = [line.strip() for line in f if line.strip()]
            self.validate_btn.setEnabled(True)
            self.status.setText(f'{len(self.domains):,} domains loaded. Click "Validate Domains" to check quality.')

    def pause_extraction(self):
        if self.worker:
            self.worker.pause()
            self.pause_btn.setEnabled(False)
            self.resume_btn.setEnabled(True)
            self.status.setText('Processing paused...')

    def resume_extraction(self):
        if self.worker:
            self.worker.resume()
            self.pause_btn.setEnabled(True)
            self.resume_btn.setEnabled(False)
            self.status.setText('Processing resumed...')

    def stop_extraction(self):
        if self.worker:
            self.worker.stop()
            self.pause_btn.setEnabled(False)
            self.resume_btn.setEnabled(False)
            self.stop_btn.setEnabled(False)
            # Enable export buttons for partial results
            if self.results:
                self.export_excel_btn.setEnabled(True)
                self.export_txt_btn.setEnabled(True)
            self.status.setText('Stopping processing...')

    def start_extraction(self):
        self.start_btn.setEnabled(False)
        self.pause_btn.setEnabled(True)
        self.resume_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.export_excel_btn.setEnabled(False)
        self.export_txt_btn.setEnabled(False)
        self.progress.setValue(0)
        self.progress.setMaximum(len(self.domains))
        self.status.setText('Starting extraction...')
        self.table.setRowCount(0)
        self.results = []
        
        # Set up auto-export path
        auto_export_path = os.path.join(os.getcwd(), f'emails_auto_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        self.worker = ExtractWorker(self.domains, auto_export_path)
        self.worker.progress.connect(self.update_progress)
        self.worker.result.connect(self.show_results)
        self.worker.status.connect(self.update_status)
        self.worker.partial_results.connect(self.update_table_live)
        self.worker.speed_update.connect(self.update_speed)
        self.worker.auto_save.connect(lambda: self.auto_save_results(auto_export_path))
        self.worker.start()

    def update_speed(self, emails_per_min, eta):
        self.speed_label.setText(f'Speed: {emails_per_min:.1f} emails/min | ETA: {eta}')

    def update_progress(self, value, maximum):
        self.progress.setValue(value)
        self.progress.setMaximum(maximum)
        self.status.setText(f'Processed {value:,}/{maximum:,} domains | Found {len(self.results):,} emails')

    def show_results(self, results, tld_counts=None):
        self.results = results
        if tld_counts:
            self.tld_counts = tld_counts
        # Force final update without timer
        self.pending_results = results
        self.process_pending_updates()
        self.export_excel_btn.setEnabled(True)
        self.export_txt_btn.setEnabled(True)
        self.start_btn.setEnabled(True)
        self.pause_btn.setEnabled(False)
        self.resume_btn.setEnabled(False)
        self.stop_btn.setEnabled(False)

    def update_status(self, text):
        self.status.setText(text)
        if 'complete' in text.lower() or 'stopped' in text.lower():
            self.start_btn.setEnabled(True)
            self.pause_btn.setEnabled(False)
            self.resume_btn.setEnabled(False)
            self.stop_btn.setEnabled(False)
            # Enable export buttons for any results (including partial)
            if self.results:
                self.export_excel_btn.setEnabled(True)
                self.export_txt_btn.setEnabled(True)

    def update_table_live(self, results):
        """Update the table with live results as they come in"""
        self.pending_results = results
        # Use timer to batch updates and prevent GUI hanging
        if not self.update_timer.isActive():
            self.update_timer.start(500)  # Update every 500ms at most

    def process_pending_updates(self):
        """Process pending table updates in batches"""
        if not self.pending_results:
            return
            
        results = self.pending_results
        self.results = results
        
        # Limit table size to prevent performance issues
        max_display_rows = 1000
        display_results = results[-max_display_rows:] if len(results) > max_display_rows else results
        
        self.table.setRowCount(len(display_results))
        
        # Update in chunks to prevent hanging
        chunk_size = 50
        for chunk_start in range(0, len(display_results), chunk_size):
            chunk_end = min(chunk_start + chunk_size, len(display_results))
            for i in range(chunk_start, chunk_end):
                if i < len(display_results):
                    row = display_results[i]
                    self.table.setItem(i, 0, QTableWidgetItem(row['domain']))
                    self.table.setItem(i, 1, QTableWidgetItem(row['email']))
                    self.table.setItem(i, 2, QTableWidgetItem(row['source_url']))
            
            # Process events between chunks to keep GUI responsive
            QApplication.processEvents()
        
        # Scroll to bottom to show latest results
        self.table.scrollToBottom()

    def auto_save_results(self, file_path):
        if self.results:
            self.save_to_excel(file_path, self.results)
            self.status.setText(f'Auto-saved {len(self.results)} results to {os.path.basename(file_path)} | Processed {self.progress.value()}/{self.progress.maximum()} domains')

    def export_to_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(self, 'Save File', '', 'Excel Files (*.xlsx)')
        if file_path:
            self.save_to_excel(file_path, self.results)
            QMessageBox.information(self, 'Export', f'Results exported to {file_path}')

    def export_to_text(self):
        file_path, _ = QFileDialog.getSaveFileName(self, 'Save File', '', 'Text Files (*.txt)')
        if file_path:
            unique_emails = set(result['email'] for result in self.results)
            with open(file_path, 'w', encoding='utf-8') as f:
                for email in sorted(unique_emails):
                    f.write(f'{email}\n')
            QMessageBox.information(self, 'Export', f'{len(unique_emails):,} unique emails exported to {file_path}')

    def save_to_excel(self, file_path, results):
        wb = Workbook()
        
        # Create Results sheet
        ws_results = wb.active
        ws_results.title = "Email Results"
        
        # Headers
        headers = ['Domain', 'Email', 'Source URL']
        for col, header in enumerate(headers, 1):
            cell = ws_results.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, result in enumerate(results, 2):
            ws_results.cell(row=row, column=1, value=result['domain'])
            ws_results.cell(row=row, column=2, value=result['email'])
            ws_results.cell(row=row, column=3, value=result['source_url'])
        
        # Auto-adjust column widths
        for column in ws_results.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_results.column_dimensions[column_letter].width = adjusted_width
        
        # Create Summary sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Summary data
        total_emails = len(results)
        unique_domains = len(set(result['domain'] for result in results))
        unique_emails = len(set(result['email'] for result in results))
        duplicate_count = total_emails - unique_emails
        
        # Domain statistics
        domain_counts = {}
        tld_counts = {}
        for result in results:
            domain = result['domain']
            domain_counts[domain] = domain_counts.get(domain, 0) + 1
            
            # Extract TLD
            tld = domain.split('.')[-1] if '.' in domain else 'unknown'
            tld_counts[tld] = tld_counts.get(tld, 0) + 1
        
        # Summary headers and data
        summary_data = [
            ['Summary Statistics', ''],
            ['Total Email Entries', total_emails],
            ['Unique Emails', unique_emails],
            ['Duplicate Entries Prevented', duplicate_count],
            ['Domains Processed', unique_domains],
            ['Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['Top Domains by Email Count', ''],
        ]
        
        # Add top 10 domains
        top_domains = sorted(domain_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        for domain, count in top_domains:
            summary_data.append([domain, count])
        
        # Add TLD distribution
        summary_data.extend([
            ['', ''],
            ['TLD Distribution', ''],
        ])
        
        top_tlds = sorted(tld_counts.items(), key=lambda x: x[1], reverse=True)[:15]
        for tld, count in top_tlds:
            summary_data.append([f'.{tld}', count])
        
        # Write summary data
        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)
            if row == 1 or row == 7:  # Headers
                ws_summary.cell(row=row, column=1).font = Font(bold=True)
        
        # Auto-adjust summary column widths
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    
    # Set modern application style
    app.setStyle('Fusion')
    
    # Global application stylesheet for consistent modern look
    app.setStyleSheet("""
        QApplication {
            font-family: 'Segoe UI', Arial, sans-serif;
        }
        QWidget {
            background-color: #ffffff;
            color: #212529;
        }
        QScrollBar:vertical {
            background-color: #f8f9fa;
            width: 12px;
            border-radius: 6px;
        }
        QScrollBar::handle:vertical {
            background-color: #6c757d;
            border-radius: 6px;
            min-height: 20px;
        }
        QScrollBar::handle:vertical:hover {
            background-color: #495057;
        }
        QScrollBar:horizontal {
            background-color: #f8f9fa;
            height: 12px;
            border-radius: 6px;
        }
        QScrollBar::handle:horizontal {
            background-color: #6c757d;
            border-radius: 6px;
            min-width: 20px;
        }
        QScrollBar::handle:horizontal:hover {
            background-color: #495057;
        }
    """)
    
    window = EmailExtractorApp()
    window.show()
    
    # Center the window on screen
    screen = app.primaryScreen().availableGeometry()
    window.move((screen.width() - window.width()) // 2, 
                (screen.height() - window.height()) // 2)
    
    sys.exit(app.exec_())
